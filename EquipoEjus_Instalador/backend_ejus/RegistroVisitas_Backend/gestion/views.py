from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from rest_framework import viewsets, generics, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.views import APIView
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from .models import Visitante
from .Serializers import VisitanteSerializer, EstadisticasSerializer
from django.db.models.functions import ExtractHour
import logging
from django.db import transaction, IntegrityError
import os
import shutil
import tempfile
import time
import subprocess
import requests
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.http import JsonResponse, HttpResponseForbidden

logger = logging.getLogger(__name__)

try:
    from openpyxl.utils import get_column_letter
    import openpyxl
    from openpyxl.styles import Font
except Exception:
    get_column_letter = None
    openpyxl = None
    Font = None

class VisitanteViewSet(viewsets.ModelViewSet):
    queryset = Visitante.objects.all()
    serializer_class = VisitanteSerializer

    def perform_create(self, serializer):
        usuario = None
        # Prefer authenticated user
        if hasattr(self.request, 'user') and self.request.user and self.request.user.is_authenticated:
            usuario = self.request.user.get_full_name() or self.request.user.username
        else:
            # Fallback to X-Usuario header if present
            usuario = self.request.headers.get('X-Usuario') or self.request.META.get('HTTP_X_USUARIO')

        # Link or create Persona by cedula so we keep a person registry
        persona = None
        cedula = serializer.validated_data.get('cedula')
        if cedula:
            from .models import Persona
            try:
                with transaction.atomic():
                    persona, created = Persona.objects.get_or_create(
                        cedula=cedula,
                        defaults={'nombre': serializer.validated_data.get('nombre', '')}
                    )
            except IntegrityError:
                try:
                    persona = Persona.objects.get(cedula=cedula)
                except Persona.DoesNotExist:
                    persona = None
            except Exception as e:
                logger.exception('Error creando Persona en perform_create: %s', e)
                persona = None

        instance = serializer.save(persona=persona)
        if usuario:
            instance.historial = (instance.historial or '') + f"\nCreado por: {usuario} - {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            instance.save()

    def perform_update(self, serializer):
        usuario = None
        if hasattr(self.request, 'user') and self.request.user and self.request.user.is_authenticated:
            usuario = self.request.user.get_full_name() or self.request.user.username
        else:
            usuario = self.request.headers.get('X-Usuario') or self.request.META.get('HTTP_X_USUARIO')

        # After updating, ensure persona is linked/updated if cedula changed
        instance = serializer.save()
        cedula = serializer.validated_data.get('cedula')
        if cedula:
            from .models import Persona
            try:
                with transaction.atomic():
                    persona, created = Persona.objects.get_or_create(
                        cedula=cedula,
                        defaults={'nombre': serializer.validated_data.get('nombre', '')}
                    )
                if instance.persona != persona:
                    instance.persona = persona
                    instance.save()
            except IntegrityError:
                try:
                    persona = Persona.objects.get(cedula=cedula)
                    if instance.persona != persona:
                        instance.persona = persona
                        instance.save()
                except Persona.DoesNotExist:
                    logger.exception('Persona esperada no encontrada tras IntegrityError')
            except Exception as e:
                logger.exception('Error creando o enlazando Persona en perform_update: %s', e)

        if usuario:
            instance.historial = (instance.historial or '') + f"\nActualizado por: {usuario} - {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            instance.save()
    
    def get_queryset(self):
        # Use select_related for persona and annotate visit_count to reduce DB queries
        queryset = Visitante.objects.select_related('persona').annotate(visit_count=Count('persona__visitas'))
        
        # Filtro manual por tipo_visita
        tipo_visita = self.request.query_params.get('tipo_visita', None)
        if tipo_visita:
            queryset = queryset.filter(tipo_visita=tipo_visita)
            
        # Filtro manual por atencion_completada
        atencion_completada = self.request.query_params.get('atencion_completada', None)
        if atencion_completada is not None:
            queryset = queryset.filter(atencion_completada=(atencion_completada.lower() == 'true'))

        # Búsqueda por nombre, cédula, teléfono, municipio o parroquia
        search = self.request.query_params.get('search') or self.request.query_params.get('q')
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(cedula__icontains=search) |
                Q(telefono__icontains=search) |
                Q(municipio__icontains=search) |
                Q(parroquia__icontains=search)
            )

        # Filtrado por referido (soporta 'NO_REFERIDO', valor específico, o 'REFERIDO' para cualquiera que no sea NO_REFERIDO)
        referir_param = self.request.query_params.get('referir_a')
        referidos_flag = self.request.query_params.get('referido')
        if referidos_flag is not None:
            # si ?referido=1 se seleccionan todos los que NO son NO_REFERIDO
            queryset = queryset.exclude(referir_a='NO_REFERIDO')
        elif referir_param:
            queryset = queryset.filter(referir_a=referir_param)
        
        # Filtrado por municipio (coincidencia parcial, insensible)
        municipio_param = self.request.query_params.get('municipio')
        if municipio_param:
            queryset = queryset.filter(municipio__icontains=municipio_param)
            
        return queryset
    
    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        hoy = timezone.now().date()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        inicio_mes = hoy.replace(day=1)
        
        total = Visitante.objects.count()
        diario = Visitante.objects.filter(fecha_hora_ingreso__date=hoy).count()
        semanal = Visitante.objects.filter(fecha_hora_ingreso__date__gte=inicio_semana).count()
        mensual = Visitante.objects.filter(fecha_hora_ingreso__date__gte=inicio_mes).count()
        
        # Visitantes activos (en sala)
        en_sala = Visitante.objects.filter(atencion_completada=False).count()
        
        data = {
            'total': total,
            'diario': diario,
            'semanal': semanal,
            'mensual': mensual,
            'enSala': en_sala,
        }

        # Devolver directamente el diccionario (no es necesario serializar aquí)
        return Response(data)
    
    @action(detail=True, methods=['post'])
    def registrar_salida(self, request, pk=None):
        visitante = self.get_object()
        visitante.fecha_hora_salida = timezone.now()
        visitante.atencion_completada = True
        visitante.save()
        
        serializer = self.get_serializer(visitante)
        return Response(serializer.data)

# ========== VISTAS PARA REPORTES ==========

class ReporteTramitesView(APIView):
    def get(self, request):
        periodo = request.GET.get('periodo', 'mes')
        # filtros opcionales desde query params
        municipio_param = request.GET.get('municipio')
        tipo_param = request.GET.get('tipo_visita')
        referir_param = request.GET.get('referir_a')
        
        # Calcular fecha de inicio según el período
        fecha_actual = timezone.now()
        if periodo == 'hoy':
            fecha_inicio = fecha_actual.replace(hour=0, minute=0, second=0, microsecond=0)
        elif periodo == 'semana':
            fecha_inicio = fecha_actual - timedelta(days=7)
        elif periodo == 'mes':
            fecha_inicio = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif periodo == 'trimestre':
            fecha_inicio = fecha_actual - timedelta(days=90)
        elif periodo == 'anio':
            fecha_inicio = fecha_actual.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            fecha_inicio = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Consulta para agrupar por tipo de trámite
        base_qs = Visitante.objects.filter(fecha_hora_ingreso__gte=fecha_inicio)
        if municipio_param:
            base_qs = base_qs.filter(municipio__icontains=municipio_param)
        if tipo_param:
            base_qs = base_qs.filter(tipo_visita=tipo_param)
        if referir_param:
            base_qs = base_qs.filter(referir_a=referir_param)

        tramites = base_qs.values('tipo_visita').annotate(
            cantidad=Count('id'),
            completados=Count('id', filter=Q(atencion_completada=True))
        ).order_by('-cantidad')
        
        # Transformar datos para el frontend
        datos = []
        for tramite in tramites:
            if tramite['cantidad'] > 0:
                porcentaje = (tramite['completados'] / tramite['cantidad']) * 100
            else: porcentaje = 0 
            datos.append({
                'nombre': tramite['tipo_visita'],
                'cantidad': tramite['cantidad'],
                'completados': tramite['completados'],
                'porcentaje_completados': round(porcentaje, 2)
            })


        
        total_tramites = sum(t['cantidad'] for t in datos)
        total_completados = sum(t['completados'] for t in datos)
        
        if total_tramites > 0:
            porcentaje_total = round((total_completados / total_tramites) * 100, 2)
        else:
            porcentaje_total = 0
        return Response({
            'success': True,
            'periodo': periodo,
            'total_tramites': sum(t['cantidad'] for t in datos),
            'porcentaje_total_completados': porcentaje_total,
            'datos': datos
        })

class ReporteVisitasMensualesView(APIView):
    def get(self, request):
        # Obtener datos de los últimos 6 meses
        fecha_actual = timezone.now()
        municipio_param = request.GET.get('municipio')
        tipo_param = request.GET.get('tipo_visita')
        referir_param = request.GET.get('referir_a')
        meses_data = []
        
        for i in range(6):
            fecha_mes = fecha_actual - timedelta(days=30*i)
            inicio_mes = fecha_mes.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # Calcular fin del mes
            if fecha_mes.month == 12:
                fin_mes = fecha_mes.replace(year=fecha_mes.year+1, month=1, day=1)
            else:
                fin_mes = fecha_mes.replace(month=fecha_mes.month+1, day=1)
            
            # Contar visitas del mes (aplicar filtros si hay)
            base_qs = Visitante.objects.filter(
                fecha_hora_ingreso__gte=inicio_mes,
                fecha_hora_ingreso__lt=fin_mes
            )
            if municipio_param:
                base_qs = base_qs.filter(municipio__icontains=municipio_param)
            if tipo_param:
                base_qs = base_qs.filter(tipo_visita=tipo_param)
            if referir_param:
                base_qs = base_qs.filter(referir_a=referir_param)

            total_visitas = base_qs.count()
            
            completados = base_qs.filter(atencion_completada=True).count()
            
            if total_visitas > 0:
                porcentaje_completados = round((completados / total_visitas) * 100, 2)
            else:
                porcentaje_completados = 0.0
            # Nombre del mes en español
            meses_es = {
                1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
                7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
            }
            
            meses_data.append({
                'mes': meses_es.get(inicio_mes.month, inicio_mes.strftime('%b')),
                'mes_numero': inicio_mes.month,
                'total_visitas': total_visitas,
                'completados': completados,
                'porcentaje_completados': porcentaje_completados
            })
        
        # Ordenar por mes (más antiguo primero)
        meses_data.reverse()
        
        return Response({
            'success': True,
            'datos': meses_data
        })

class ReporteTendenciaSemanalView(APIView):
    def get(self, request):
        # Obtener datos de las últimas 6 semanas
        fecha_actual = timezone.now()
        municipio_param = request.GET.get('municipio')
        tipo_param = request.GET.get('tipo_visita')
        referir_param = request.GET.get('referir_a')
        semanas_data = []
        
        for i in range(6):
            inicio_semana = fecha_actual - timedelta(weeks=i, days=fecha_actual.weekday())
            fin_semana = inicio_semana + timedelta(days=7)
            
            base_qs = Visitante.objects.filter(
                fecha_hora_ingreso__gte=inicio_semana,
                fecha_hora_ingreso__lt=fin_semana
            )
            if municipio_param:
                base_qs = base_qs.filter(municipio__icontains=municipio_param)
            if tipo_param:
                base_qs = base_qs.filter(tipo_visita=tipo_param)
            if referir_param:
                base_qs = base_qs.filter(referir_a=referir_param)

            # Contar visitas de la semana
            total_visitas = base_qs.count()
            
            # Calcular promedio diario
            promedio_diario = round(total_visitas / 7, 2) if total_visitas > 0 else 0
            
            semanas_data.append({
                'semana': f"Sem {6-i}",
                'semana_numero': 6 - i,
                'total_visitas': total_visitas,
                'promedio_diario': promedio_diario,
                'fecha_inicio': inicio_semana.strftime('%d/%m'),
                'fecha_fin': fin_semana.strftime('%d/%m')
            })
        
        # Ordenar por semana (más antigua primero)
        semanas_data.reverse()
        
        return Response({
            'success': True,
            'datos': semanas_data
        })


class ReporteDiarioView(APIView):
    """Devuelve conteo de visitas por día para los últimos N días (por defecto 14)."""
    def get(self, request):
        try:
            dias = int(request.GET.get('days', 14))
        except ValueError:
            dias = 14

        fecha_actual = timezone.now().date()
        datos = []

        municipio_param = request.GET.get('municipio')
        tipo_param = request.GET.get('tipo_visita')
        referir_param = request.GET.get('referir_a')

        for i in range(dias - 1, -1, -1):
            dia = fecha_actual - timedelta(days=i)
            base_qs = Visitante.objects.filter(fecha_hora_ingreso__date=dia)
            if municipio_param:
                base_qs = base_qs.filter(municipio__icontains=municipio_param)
            if tipo_param:
                base_qs = base_qs.filter(tipo_visita=tipo_param)
            if referir_param:
                base_qs = base_qs.filter(referir_a=referir_param)

            total = base_qs.count()
            datos.append({
                'fecha': dia.strftime('%Y-%m-%d'),
                'label': dia.strftime('%d/%m'),
                'visitas': total
            })

        return Response({
            'success': True,
            'dias': dias,
            'datos': datos
        })

class ReporteEstadisticasView(APIView):
    def get(self, request):
        fecha_actual = timezone.now()

        # aceptar filtros opcionales para estadísticas
        municipio_param = request.GET.get('municipio')
        tipo_param = request.GET.get('tipo_visita')
        referir_param = request.GET.get('referir_a')

        # Base queryset con filtros aplicados
        base_qs = Visitante.objects.all()
        if municipio_param:
            base_qs = base_qs.filter(municipio__icontains=municipio_param)
        if tipo_param:
            base_qs = base_qs.filter(tipo_visita=tipo_param)
        if referir_param:
            base_qs = base_qs.filter(referir_a=referir_param)

        # 1. CALCULAR PROMEDIO DESDE PRIMERA VISITA (sobre la muestra filtrada)
        primera_visita = base_qs.order_by('fecha_hora_ingreso').first()
        if not primera_visita:
            return Response({
                'success': True,
                'total_visitas': 0,
                'promedio_diario': 0,
                'tramite_mas_comun': 'No disponible',
                'porcentaje_completados': 0,
                'municipio_mas_visitado': 'No disponible',
                'mensaje': 'No hay visitas registradas en el periodo/filtro seleccionado'
            })

        primera_fecha = primera_visita.fecha_hora_ingreso.date()
        hoy = fecha_actual.date()
        dias_transcurridos = (hoy - primera_fecha).days + 1
        if dias_transcurridos < 1:
            dias_transcurridos = 1

        total_visitas = base_qs.count()
        promedio_diario = total_visitas / dias_transcurridos if dias_transcurridos > 0 else 0

        tramite_mas_comun_data = base_qs.values('tipo_visita').annotate(count=Count('id')).order_by('-count').first()
        tramite_mas_comun = tramite_mas_comun_data['tipo_visita'] if tramite_mas_comun_data else 'No disponible'

        completados_total = base_qs.filter(atencion_completada=True).count()
        porcentaje_completados = round((completados_total / total_visitas) * 100, 2) if total_visitas > 0 else 0.0

        try:
            municipio_mas_visitado_data = base_qs.exclude(municipio__isnull=True).exclude(municipio='').values('municipio').annotate(count=Count('id')).order_by('-count').first()
            municipio = municipio_mas_visitado_data['municipio'] if municipio_mas_visitado_data else 'No disponible'
        except:
            municipio = 'No disponible'

        
        try:
            parroquia_mas_visitada_data = base_qs.exclude(parroquia__isnull=True).exclude(parroquia='').values('parroquia').annotate(count=Count('id')).order_by('-count').first()
            parroquia = parroquia_mas_visitada_data['parroquia'] if parroquia_mas_visitada_data else 'No disponible'
        except:
            parroquia = 'No disponible'

        tipos_tramite_count = base_qs.values('tipo_visita').distinct().count()
        activos = base_qs.filter(atencion_completada=False).count()

        return Response({
            'success': True,
            'total_visitas': total_visitas,
            'visitas_completadas': completados_total,
            'visitas_pendientes': total_visitas - completados_total,
            'visitas_activas': activos,
            'promedio_diario': round(promedio_diario, 2),
            'tramite_mas_comun': tramite_mas_comun,
            'porcentaje_completados': porcentaje_completados,
            'municipio_mas_visitado': municipio,
            'parroquia_mas_visitada': parroquia,
            'tipos_tramite_diferentes': tipos_tramite_count,
            'dias_transcurridos': dias_transcurridos,
            'primera_visita_fecha': primera_fecha.strftime('%d/%m/%Y'),
            'hoy': hoy.strftime('%d/%m/%Y'),
            '_formula_promedio': f"{total_visitas} visitas / {dias_transcurridos} días = {round(promedio_diario, 2)}",
            '_muestra_tramites': base_qs.values('tipo_visita').annotate(count=Count('id')).order_by('-count')[:5]
        })

class ExportarReportePDFView(APIView):
    def get(self, request):
        periodo = request.GET.get('periodo', 'mes')

        # Preparar respuesta
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{periodo}_{datetime.now().strftime("%Y%m%d")}.pdf"'

        buffer = BytesIO()

        # Usar Platypus para estructura profesional
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=40, leftMargin=40,
                                topMargin=60, bottomMargin=40)
        styles = getSampleStyleSheet()
        styleH = ParagraphStyle('Heading', parent=styles['Heading1'], fontSize=14, leading=18)
        styleN = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10, leading=14)

        elements = []

        # Portada con líneas exactas solicitadas
        cover_style = ParagraphStyle('Cover', parent=styles['Title'], alignment=1, fontSize=14, leading=18)
        cover_lines = [
            'REPUBLICA BOLIVARIANA DE VENEZUELA',
            'TRIBUNAL SUPREMO DE JUSTICIA',
            'DIRECCION EJECUTIVA DE LA MAGISTRATURA',
            'EQUIPO DE JUSTICIA SOCIAL',
            'BARQUISIMETO-ESTADO LARA'
        ]
        elements.append(Spacer(1, 40))
        for line in cover_lines:
            elements.append(Paragraph(line, cover_style))
            elements.append(Spacer(1, 6))
        # Subtítulo con período y fecha de generación
        subtitle = Paragraph(f'Reporte de Visitas - Período: {periodo}', ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=1, fontSize=11))
        genDate = Paragraph(f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ParagraphStyle('GenDate', parent=styles['Normal'], alignment=1, fontSize=9))
        # Detectar usuario que generó el reporte
        usuario = None
        if hasattr(request, 'user') and request.user and request.user.is_authenticated:
            usuario = request.user.get_full_name() or request.user.username
        else:
            usuario = request.headers.get('X-Usuario') or request.META.get('HTTP_X_USUARIO')
        genUser = Paragraph(f'Generado por: {usuario}' if usuario else 'Generado por: Usuario no identificado', ParagraphStyle('GenUser', parent=styles['Normal'], alignment=1, fontSize=9))
        elements.extend([Spacer(1, 12), subtitle, Spacer(1, 6), genDate, Spacer(1,4), genUser, PageBreak()])

        # En lugar de tabla de contenidos, colocamos directamente las secciones
        # Todas las tablas se agregarán en la misma hoja (sin saltos de página entre ellas)

        # Estadísticas principales (diseño más profesional)
        elements.append(Paragraph('Estadísticas Principales', styleH))
        stats_view = ReporteEstadisticasView()
        stats_response = stats_view.get(request)
        stats_data = stats_response.data if getattr(stats_response, 'data', None) is not None else {}

        stats_table_data = []
        stats_rows = [
            ('Total visitas', stats_data.get('total_visitas', 0)),
            ('Promedio diario', stats_data.get('promedio_diario', 0)),
            ('Trámite más común', stats_data.get('tramite_mas_comun', 'N/A')),
            ('Porcentaje completados', f"{stats_data.get('porcentaje_completados', 0)}%"),
            ('Municipio más visitado', stats_data.get('municipio_mas_visitado', 'N/A')),
        ]

        # Crear filas con clave en negrita y valor alineado a la izquierda
        for k, v in stats_rows:
            stats_table_data.append([Paragraph(f'<b>{k}</b>', styleN), Paragraph(str(v), styleN)])

        t = Table(stats_table_data, colWidths=[240, 260], hAlign='LEFT')
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F3F6FA')),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#D6DCE6')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.extend([Spacer(1,8), t, Spacer(1,12)])

        # Distribución por trámite
        elements.append(Paragraph('Distribución por Trámite', styleH))
        tramites_view = ReporteTramitesView()
        tramites_response = tramites_view.get(request)
        tramites_data = tramites_response.data if getattr(tramites_response, 'data', None) is not None else {'datos': []}
        tramites_table = [[ 'Trámite', 'Cantidad', 'Completados', '% Completados' ]]
        for tr in tramites_data.get('datos', []):
            tramites_table.append([ tr.get('nombre',''), tr.get('cantidad',0), tr.get('completados',0), tr.get('porcentaje_completados','') ])
        if len(tramites_table) == 1:
            tramites_table.append(['No hay datos', '', '', ''])

        tt = Table(tramites_table, colWidths=[260, 80, 80, 80], hAlign='LEFT')
        tt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C7857')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#D6DCE6')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN',(1,1),(-1,-1),'CENTER'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.extend([Spacer(1,6), tt, Spacer(1,12)])

        # Visitas mensuales
        elements.append(Paragraph('Visitas Mensuales', styleH))
        mensual_view = ReporteVisitasMensualesView()
        mensual_response = mensual_view.get(request)
        mensual_data = mensual_response.data if getattr(mensual_response, 'data', None) is not None else {'datos': []}
        mensual_table = [[ 'Mes', 'Visitas', 'Completados' ]]
        for m in mensual_data.get('datos', []):
            mensual_table.append([ m.get('mes',''), m.get('total_visitas', m.get('visitas',0)), m.get('completados',0) ])
        if len(mensual_table) == 1:
            mensual_table.append(['No hay datos', '', ''])

        mt = Table(mensual_table, colWidths=[200, 160, 160], hAlign='LEFT')
        mt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2B6CE4')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#D6DCE6')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN',(1,1),(-1,-1),'CENTER'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.extend([Spacer(1,6), mt, Spacer(1,12)])

        # Agregar firmas al final de la misma hoja (dos bloques lado a lado)
        elements.append(Spacer(1, 18))
        sign_table_data = [
            ['_______________________________', '_______________________________'],
            ['Coordinadora de Equipo de Justicia Social', 'Directora Administrativa Regional']
        ]
        sign_table = Table(sign_table_data, colWidths=[260, 260], hAlign='CENTER')
        sign_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 11),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(sign_table)
        elements.append(Spacer(1, 18))
        elements.append(Paragraph('Fin del reporte', styleN))

        # Build PDF
        try:
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            return response
        except Exception as e:
            # Fallback: return simple text response if generation fails
            logger.exception('Error generando PDF: %s', e)
            buffer.close()
            return HttpResponse(f"Error generando PDF: {e}", status=500)

class ExportarReporteExcelView(APIView):
    def get(self, request):
        periodo = request.GET.get('periodo', 'mes')
        
        # Obtener datos
        tramites_view = ReporteTramitesView()
        tramites_response = tramites_view.get(request)
        tramites_data = tramites_response.data if tramites_response.status_code == 200 else {'datos': []}
        
        mensual_view = ReporteVisitasMensualesView()
        mensual_response = mensual_view.get(request)
        mensual_data = mensual_response.data if mensual_response.status_code == 200 else {'datos': []}
        
        stats_view = ReporteEstadisticasView()
        stats_response = stats_view.get(request)
        stats_data = stats_response.data if stats_response.status_code == 200 else {}
        
        # Crear DataFrames de pandas
        df_tramites = pd.DataFrame(tramites_data.get('datos', []))
        df_mensual = pd.DataFrame(mensual_data.get('datos', []))
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Hoja de resumen
                df_resumen = pd.DataFrame([{
                    'Total Visitas': stats_data.get('total_visitas', 0),
                    'Promedio Diario': stats_data.get('promedio_diario', 0),
                    'Trámite más Común': stats_data.get('tramite_mas_comun', ''),
                    'Porcentaje Completados': f"{stats_data.get('porcentaje_completados', 0)}%",
                    'Municipio más Visitado': stats_data.get('municipio_mas_visitado', ''),
                    # 'Hora Pico' removed per request
                    'Período Reporte': periodo
                }])
                df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Hoja de trámites
                if not df_tramites.empty:
                    df_tramites.to_excel(writer, sheet_name='Trámites', index=False)
                
                # Hoja mensual
                if not df_mensual.empty:
                    df_mensual.to_excel(writer, sheet_name='Visitas Mensuales', index=False)
                
                # Hoja de metadata
                df_metadata = pd.DataFrame([{
                    'Período': periodo,
                    'Fecha Generación': datetime.now().strftime('%d/%m/%Y %H:%M'),
                    'Generado por': (request.user.get_full_name() or request.user.username) if hasattr(request, 'user') and request.user.is_authenticated else (request.headers.get('X-Usuario') or request.META.get('HTTP_X_USUARIO') or ''),
                    'Total Registros': len(df_tramites) + len(df_mensual)
                }])
                df_metadata.to_excel(writer, sheet_name='Metadata', index=False)
                # Hoja de firmas (en Excel para impresión)
                df_firmas = pd.DataFrame({
                    'Coordinadora': [''],
                    'Directora Administrativa Regional': ['']
                })
                df_firmas.to_excel(writer, sheet_name='Firmas', index=False)

                # Mejoras de formato si openpyxl está disponible
                try:
                    if openpyxl is not None and get_column_letter is not None:
                        wb = writer.book
                        # Formato resumen
                        ws_res = writer.sheets.get('Resumen') or wb['Resumen']
                        # Ajustar anchos de columna
                        for i, col in enumerate(df_resumen.columns, 1):
                            ws_res.column_dimensions[get_column_letter(i)].width = max(15, len(col) + 6)
                        # Encabezados en negrita
                        if Font is not None:
                            for cell in ws_res[1]:
                                cell.font = Font(bold=True)

                        # Formato trámites
                        if not df_tramites.empty:
                            ws_tr = writer.sheets.get('Trámites') or wb['Trámites']
                            for i, col in enumerate(df_tramites.columns, 1):
                                ws_tr.column_dimensions[get_column_letter(i)].width = max(12, len(str(col)) + 6)
                            for cell in ws_tr[1]:
                                cell.font = Font(bold=True)

                        # Formato mensual
                        if not df_mensual.empty:
                            ws_m = writer.sheets.get('Visitas Mensuales') or wb['Visitas Mensuales']
                            for i, col in enumerate(df_mensual.columns, 1):
                                ws_m.column_dimensions[get_column_letter(i)].width = max(12, len(str(col)) + 6)
                            for cell in ws_m[1]:
                                cell.font = Font(bold=True)

                        # Hoja firmas: ajustar ancho y centrar texto
                        ws_f = writer.sheets.get('Firmas') or wb['Firmas']
                        ws_f.column_dimensions[get_column_letter(1)].width = 40
                        ws_f.column_dimensions[get_column_letter(2)].width = 40
                        for cell in ws_f[1]:
                            cell.font = Font(bold=False)
                        # Poner líneas de firma como texto grande
                        ws_f.cell(row=1, column=1).value = '_______________________________'
                        ws_f.cell(row=1, column=2).value = '_______________________________'
                        ws_f.cell(row=2, column=1).value = 'Coordinadora de Equipo de Justicia Social'
                        ws_f.cell(row=2, column=2).value = 'Directora Administrativa Regional'
                except Exception as fe:
                    # Si falla el formateo con openpyxl, seguir sin formato
                    logger.warning('No se pudo aplicar formato Excel (openpyxl): %s', fe)
            
            output.seek(0)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"reporte_{periodo}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            
        except Exception as e:
            # Fallback simple si hay error con pandas
            output = BytesIO()
            output.write(b"Reporte generado el " + datetime.now().strftime('%d/%m/%Y %H:%M').encode())
            output.write(b"\nPeriodo: " + periodo.encode())
            output.seek(0)
            content_type = 'text/plain'
            filename = f"reporte_{periodo}_{datetime.now().strftime('%Y%m%d')}.txt"
        
        # Preparar respuesta
        response = HttpResponse(output.getvalue(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    # ========== VISTA PARA REPORTE DE REFERIDOS ==========

class ReporteReferidosView(APIView):
    def get(self, request):
        periodo = request.GET.get('periodo', 'mes')
        
        # Calcular fecha de inicio según el período
        fecha_actual = timezone.now()
        if periodo == 'hoy':
            fecha_inicio = fecha_actual.replace(hour=0, minute=0, second=0, microsecond=0)
        elif periodo == 'semana':
            fecha_inicio = fecha_actual - timedelta(days=7)
        elif periodo == 'mes':
            fecha_inicio = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif periodo == 'trimestre':
            fecha_inicio = fecha_actual - timedelta(days=90)
        elif periodo == 'anio':
            fecha_inicio = fecha_actual.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            fecha_inicio = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # filtros opcionales
        municipio_param = request.GET.get('municipio')
        tipo_param = request.GET.get('tipo_visita')
        referir_param = request.GET.get('referir_a')

        # 1. Obtener total de visitantes en el período (aplicar filtros si existen)
        base_total = Visitante.objects.filter(fecha_hora_ingreso__gte=fecha_inicio)
        if municipio_param:
            base_total = base_total.filter(municipio__icontains=municipio_param)
        if tipo_param:
            base_total = base_total.filter(tipo_visita=tipo_param)
        if referir_param:
            base_total = base_total.filter(referir_a=referir_param)

        total_visitantes = base_total.count()

        # 2. Obtener visitantes referidos (excluyendo 'NO_REFERIDO')
        visitantes_referidos = Visitante.objects.filter(fecha_hora_ingreso__gte=fecha_inicio).exclude(referir_a='NO_REFERIDO')
        if municipio_param:
            visitantes_referidos = visitantes_referidos.filter(municipio__icontains=municipio_param)
        if tipo_param:
            visitantes_referidos = visitantes_referidos.filter(tipo_visita=tipo_param)
        if referir_param:
            visitantes_referidos = visitantes_referidos.filter(referir_a=referir_param)
        
        total_referidos = visitantes_referidos.count()
        
        # 3. Calcular porcentaje de referidos
        if total_visitantes > 0:
            porcentaje_referidos = round((total_referidos / total_visitantes) * 100, 2)
        else:
            porcentaje_referidos = 0.0
        
        # 4. Agrupar por institución
        instituciones_data = visitantes_referidos.values('referir_a').annotate(
            total_referidos=Count('id'),
            tramites_comunes=Count('tipo_visita', distinct=True)
        ).order_by('-total_referidos')
        
        # 5. Transformar datos para el frontend
        instituciones_formateadas = []
        for inst in instituciones_data:
            # Obtener nombre legible de la institución
            nombre_institucion = dict(Visitante.INSTITUCION_CHOICES).get(inst['referir_a'], inst['referir_a'])
            
            # Para OTRA_INSTITUCION, obtener el valor específico
            if inst['referir_a'] == 'OTRA_INSTITUCION':
                # Obtener las otras instituciones especificadas
                otras_instituciones = visitantes_referidos.filter(
                    referir_a='OTRA_INSTITUCION'
                ).values('otra_institucion').annotate(
                    count=Count('id')
                ).order_by('-count')[:3]
                
                nombre_institucion = "Otras Instituciones"
                tramites_list = [item['otra_institucion'] for item in otras_instituciones if item['otra_institucion']]
            else:
                # Obtener trámites más comunes para esta institución
                tramites_comunes = visitantes_referidos.filter(
                    referir_a=inst['referir_a']
                ).values('tipo_visita').annotate(
                    count=Count('id')
                ).order_by('-count')[:3]
                
                tramites_list = [dict(Visitante.TIPO_VISITA_CHOICES).get(t['tipo_visita'], t['tipo_visita']) 
                               for t in tramites_comunes]
            
            if inst['total_referidos'] > 0:
                porcentaje = round((inst['total_referidos'] / total_referidos) * 100, 2) if total_referidos > 0 else 0
            else:
                porcentaje = 0.0
            
            instituciones_formateadas.append({
                'institucion': nombre_institucion,
                'codigo_institucion': inst['referir_a'],
                'total_referidos': inst['total_referidos'],
                'porcentaje': porcentaje,
                'tramites_comunes': tramites_list[:3]  # Máximo 3 trámites
            })
        
        # 6. Agregar estadísticas generales
        metadata = {
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d %H:%M:%S'),
            'fecha_fin': fecha_actual.strftime('%Y-%m-%d %H:%M:%S'),
            'periodo_seleccionado': periodo,
            'generado_en': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return Response({
            'success': True,
            'periodo': periodo,
            'total_visitantes': total_visitantes,
            'total_referidos': total_referidos,
            'porcentaje_referidos': porcentaje_referidos,
            'instituciones': instituciones_formateadas,
            'metadata': metadata
        })

# ========== VISTA PARA ESTADÍSTICAS DE REFERIDOS ==========

class EstadisticasReferidosView(APIView):
    def get(self, request):
        """Estadísticas específicas de referidos para dashboard"""
        fecha_actual = timezone.now()
        
        # Estadísticas de hoy
        inicio_hoy = fecha_actual.replace(hour=0, minute=0, second=0, microsecond=0)
        referidos_hoy = Visitante.objects.filter(
            fecha_hora_ingreso__gte=inicio_hoy
        ).exclude(referir_a='NO_REFERIDO').count()
        
        total_hoy = Visitante.objects.filter(
            fecha_hora_ingreso__gte=inicio_hoy
        ).count()
        
        # Estadísticas de la semana
        inicio_semana = fecha_actual - timedelta(days=7)
        referidos_semana = Visitante.objects.filter(fecha_hora_ingreso__gte=inicio_semana).exclude(referir_a='NO_REFERIDO').count()
        
        total_semana = Visitante.objects.filter(
            fecha_hora_ingreso__gte=inicio_semana
        ).count()
        
        # Estadísticas del mes
        inicio_mes = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        referidos_mes = Visitante.objects.filter(fecha_hora_ingreso__gte=inicio_mes).exclude(referir_a='NO_REFERIDO').count()
        
        total_mes = Visitante.objects.filter(
            fecha_hora_ingreso__gte=inicio_mes
        ).count()
        
        # Instituciones más frecuentes (top 5)
        instituciones_top = Visitante.objects.exclude(referir_a='NO_REFERIDO').values('referir_a').annotate(
            total=Count('id')
        ).order_by('-total')[:5]
        
        instituciones_top_formateadas = []
        for inst in instituciones_top:
            nombre = dict(Visitante.INSTITUCION_CHOICES).get(inst['referir_a'], inst['referir_a'])
            instituciones_top_formateadas.append({
                'institucion': nombre,
                'total': inst['total'],
                'codigo': inst['referir_a']
            })
        
        # Calcular porcentajes
        porcentaje_hoy = round((referidos_hoy / total_hoy * 100), 2) if total_hoy > 0 else 0
        porcentaje_semana = round((referidos_semana / total_semana * 100), 2) if total_semana > 0 else 0
        porcentaje_mes = round((referidos_mes / total_mes * 100), 2) if total_mes > 0 else 0
        
        return Response({
            'success': True,
            'estadisticas': {
                'hoy': {
                    'referidos': referidos_hoy,
                    'total': total_hoy,
                    'porcentaje': porcentaje_hoy
                },
                'semana': {
                    'referidos': referidos_semana,
                    'total': total_semana,
                    'porcentaje': porcentaje_semana
                },
                'mes': {
                    'referidos': referidos_mes,
                    'total': total_mes,
                    'porcentaje': porcentaje_mes
                },
                'total_general': Visitante.objects.filter(referir_a__ne='NO_REFERIDO').count(),
                'instituciones_top': instituciones_top_formateadas
            },
            'periodo_actual': fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
        })
    
# ========== VISTAS PARA OPCIONES ==========

class OpcionesTipoVisitaView(APIView):
    """Vista para obtener las opciones de tipo de visita"""
    def get(self, request):
        return Response(Visitante.TIPO_VISITA_CHOICES)

class OpcionesInstitucionesView(APIView):
    """Vista para obtener las opciones de instituciones"""
    def get(self, request):
        return Response(Visitante.INSTITUCION_CHOICES)


class OpcionesMunicipiosView(APIView):
    """Vista para obtener la lista de municipios únicos"""
    def get(self, request):
        municipios = Visitante.objects.exclude(municipio__isnull=True).exclude(municipio='').values_list('municipio', flat=True).distinct().order_by('municipio')
        return Response(list(municipios))


class LoginView(APIView):
    """Login sencillo que devuelve un token DRF"""
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        if not username or not password:
            return Response({'detail': 'Username and password required'}, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(request, username=username, password=password)
        if user is None:
            return Response({'detail': 'Credenciales inválidas'}, status=status.HTTP_400_BAD_REQUEST)

        token, created = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'username': user.get_full_name() or user.username,
            'id': user.id
        })


class LogoutView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        # eliminar token para forzar logout
        try:
            Token.objects.filter(user=request.user).delete()
        except Exception as e:
            logger.exception('Error eliminando token en LogoutView: %s', e)
        return Response(status=status.HTTP_204_NO_CONTENT)


class CurrentUserView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({'username': user.get_full_name() or user.username, 'id': user.id})


class RegisterView(APIView):
    """Registro público de usuarios (crea usuario y devuelve token)."""
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        full_name = request.data.get('full_name') or ''

        if not username or not password:
            return Response({'detail': 'username and password are required'}, status=status.HTTP_400_BAD_REQUEST)

        from django.contrib.auth import get_user_model
        User = get_user_model()
        if User.objects.filter(username=username).exists():
            return Response({'detail': 'username already exists'}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(username=username, password=password)
        # optionally set full name if provided
        if full_name:
            try:
                user.first_name = full_name
                user.save()
            except Exception as e:
                logger.exception('Error guardando full_name en RegisterView: %s', e)

        token, _ = Token.objects.get_or_create(user=user)
        return Response({'token': token.key, 'username': user.get_full_name() or user.username, 'id': user.id})


# ========== ENDPOINTS DE ACTUALIZACIÓN / VERSION ==========
def _is_local_request(request):
    addr = request.META.get('REMOTE_ADDR')
    return addr in ('127.0.0.1', '::1')


@require_GET
def version(request):
    """Devuelve la versión actual del instalador/app (archivo VERSION en BASE_DIR)."""
    vf = os.path.join(settings.BASE_DIR, 'VERSION')
    v = '0.0.0'
    try:
        if os.path.exists(vf):
            with open(vf, 'r') as f:
                v = f.read().strip()
    except Exception:
        v = '0.0.0'
    return JsonResponse({'version': v})


@csrf_exempt
@require_POST
def update(request):
    """Realiza update offline: backup DB, descarga instalador, ejecuta, restaura si falla, ejecuta migraciones."""
    # Seguridad: solo localhost o token correcto
    token = request.headers.get('X-UPDATE-TOKEN') or request.POST.get('token')
    configured = getattr(settings, 'UPDATE_TOKEN', None)
    if configured and token != configured:
        return HttpResponseForbidden('invalid token')

    if not _is_local_request(request) and not configured:
        return HttpResponseForbidden('forbidden')

    installer_url = request.POST.get('url') or getattr(settings, 'UPDATE_INSTALLER_URL', None)
    app_dir = settings.BASE_DIR
    db_path = os.path.join(app_dir, 'db.sqlite3')
    backup_path = db_path + f'.bak.{int(time.time())}'

    try:
        # Stop service if exists
        try:
            subprocess.run(['sc', 'stop', 'EquipoEjus_Backend'], check=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except Exception:
            pass

        # Backup DB
        if os.path.exists(db_path):
            shutil.copy2(db_path, backup_path)

        # Download installer to temp if URL provided
        tmp_inst = None
        if installer_url:
            fd, tmp_inst = tempfile.mkstemp(suffix='.exe')
            os.close(fd)
            r = requests.get(installer_url, stream=True, timeout=60)
            r.raise_for_status()
            with open(tmp_inst, 'wb') as fh:
                for chunk in r.iter_content(8192):
                    fh.write(chunk)

        # If installer provided as upload (multipart), save it
        if 'installer' in request.FILES:
            fd, tmp2 = tempfile.mkstemp(suffix='.exe')
            os.close(fd)
            with open(tmp2, 'wb') as fh:
                for chunk in request.FILES['installer'].chunks():
                    fh.write(chunk)
            tmp_inst = tmp2

        # Execute installer silently if present
        exitcode = None
        if tmp_inst and os.path.exists(tmp_inst):
            args = [tmp_inst, '/VERYSILENT', '/SUPPRESSMSGBOXES', '/NORESTART']
            proc = subprocess.run(args)
            exitcode = proc.returncode

        # Run migrations
        manage_py = os.path.join(app_dir, 'manage.py')
        python_exe = os.path.join(app_dir, 'venv', 'Scripts', 'python.exe')
        if os.path.exists(python_exe):
            subprocess.run([python_exe, manage_py, 'migrate', '--noinput'], check=False)

        # Start service
        try:
            subprocess.run(['sc', 'start', 'EquipoEjus_Backend'], check=False)
        except Exception:
            pass

        return JsonResponse({'ok': True, 'installer_exitcode': exitcode})
    except Exception as e:
        # Restore DB on error
        if os.path.exists(backup_path):
            try:
                shutil.copy2(backup_path, db_path)
            except Exception:
                pass
        return JsonResponse({'error': str(e)}, status=500)